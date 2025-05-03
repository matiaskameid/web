from django.contrib.auth.decorators import login_required

import os
import re
import json
import zipfile
import base64
from io import BytesIO

import pandas as pd
from django import forms
from django.shortcuts      import render
from django.conf           import settings
from django.contrib        import messages
from django.http           import HttpResponse
from django.core.files.storage import default_storage
from django.core.files.base    import ContentFile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles       import Alignment, Font, Border, Side
from openpyxl.utils        import get_column_letter

from .forms import UploadFileForm, ContactInfoForm

# alias para el formset
ContactInfoFormSet = forms.formset_factory(ContactInfoForm, extra=0)

# rutas dentro de la app
APP_DIR       = os.path.join(settings.BASE_DIR, 'consignaciones_atico')
CONTACTS_FILE = os.path.join(APP_DIR, 'contact_data.json')
LOGO_PATH     = os.path.join(APP_DIR, 'static', 'consignaciones_atico', 'logo.png')


def load_contact_data():
    try:
        with open(CONTACTS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}


def save_contact_data(data):
    try:
        with open(CONTACTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
    except Exception as e:
        messages.error(None, f"Error guardando contactos: {e}")


def load_logo_bytes():
    try:
        with open(LOGO_PATH, 'rb') as f:
            return f.read()
    except:
        return None


def extract_editoriales(file_bytes):
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=5)
    # forzamos todas las columnas a string
    df.columns = df.columns.map(str).str.strip()
    edits = []
    for c in df.columns:
        if re.search(r'consignacion', c, re.IGNORECASE):
            ed = re.sub(r'(?i)consignacion(es)?', '', c)
            ed = re.sub(r'\s+', ' ', ed)
            ed = re.sub(r'[:]+', '', ed)
            ed = re.sub(r'[0-9-]+', '', ed)
            edits.append(ed.strip().upper() or 'SIN EDITORIAL')
    return sorted(set(edits))


def create_export_excel(df, editorial, logo_content=None, contact_info=None):
    # … copiar aquí tu implementación de OpenPyXL exactamente …
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación"
    ws.sheet_view.showGridLines = False

    title_font  = Font(name="Arial", size=16, bold=True)
    header_font = Font(name="Arial", size=11, bold=True)
    normal_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws.row_dimensions[1].height = 45
    if logo_content:
        try:
            img = OpenpyxlImage(BytesIO(logo_content))
            img.width, img.height = 80, 50
            ws.add_image(img, "A1")
        except:
            pass

    ws.merge_cells("B1:D2")
    cell_title = ws["B1"]
    titulo = f"LIQUIDACION CONSIGNACIONES {editorial}"
    cell_title.value     = titulo
    cell_title.font      = title_font
    cell_title.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B3:D6")
    cell_cli = ws["B3"]
    cell_cli.value = (
        "CLIENTE: Librería Virtual y Distribuidora El Ático Ltda.\n"
        "Venta y Distribución de Libros\n"
        "General Bari 234, Providencia - Santiago, Teléfono: (56)2 21452308\n"
        "Rut: 76082908-0"
    )
    cell_cli.font      = normal_font
    cell_cli.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    fields = ["PROVEEDOR:", "CONTACTO:", "FONO / MAIL:", "DESCUENTO:", "PAGO:", "FECHA:"]
    for i, field in enumerate(fields, start=8):
        ws.cell(row=i, column=2, value=field).font = header_font
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        c = ws.cell(row=i, column=3)
        key = field.replace(":", "")
        c.value = contact_info.get(key, "") if contact_info else ""
        c.font  = normal_font
        for col in (2, 3, 4):
            ws.cell(row=i, column=col).border = thin_border

    start_row, start_col = 16, 2
    for off, header in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col+off, value=header)
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border

    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row+1):
        for off, val in enumerate(row):
            c = ws.cell(row=r_idx, column=start_col+off, value=val)
            if off == 2:
                try:    c.value = int(val)
                except: pass
                c.number_format = "0"
            c.font      = normal_font
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = thin_border

    unidades_w = len("Unidades a liquidar")+2
    prod_w     = max(df["Producto"].astype(str).map(len).max() if not df.empty else len("Producto"), len("Producto"))+5
    isbn_w     = 15
    required = 2*len(titulo)
    current  = unidades_w + prod_w + isbn_w
    if required>current:
        prod_w += (required-current)

    ws.column_dimensions[get_column_letter(2)].width = unidades_w
    ws.column_dimensions[get_column_letter(3)].width = prod_w
    ws.column_dimensions[get_column_letter(4)].width = isbn_w

    max_c = ws.max_column
    if max_c > 5:
        ws.delete_cols(6, max_c-5)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def process_master_file(file_bytes, logo_content=None, contact_infos=None):
    # 1) leer y forzar columnas a string
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=5)
    df.columns = df.columns.map(str).str.strip()
    if "Código" in df.columns:
        df.rename(columns={"Código": "Codigo"}, inplace=True)

    output_files = {}
    no_data      = []

    for col in df.columns:
        if re.search(r'consignacion', col, re.IGNORECASE):
            name = re.sub(r'(?i)consignacion(es)?','',col)
            name = re.sub(r'\s+',' ',name)
            name = re.sub(r'[:]+','',name)
            name = re.sub(r'[0-9-]+','',name)
            name = (name.strip().upper() or "SIN EDITORIAL")

            if not all(x in df.columns for x in ["Producto", "Codigo", "BODEGA GENERAL BARI"]):
                return {}, []

            temp = df[["Producto", "Codigo", "BODEGA GENERAL BARI", col]].copy()
            temp.rename(columns={col: "Consignaciones"}, inplace=True)
            temp = temp[temp["BODEGA GENERAL BARI"] >= 0]
            temp["Unidades a liquidar"] = temp["Consignaciones"] - temp["BODEGA GENERAL BARI"]
            temp = temp[temp["Unidades a liquidar"] > 0].sort_values("Producto")

            if temp.empty:
                no_data.append(name)
            else:
                export_df = temp[["Unidades a liquidar","Producto","Codigo"]].copy()
                export_df.rename(columns={"Codigo":"ISBN"}, inplace=True)
                export_df["ISBN"] = export_df["ISBN"].astype(str).apply(lambda x: x.split("/")[0][:13])
                ci = contact_infos.get(name, {}) if contact_infos else {}
                xlsx_bytes = create_export_excel(export_df, name, logo_content, ci)
                filename   = f"Liquidacion_Consignaciones_{name}.xlsx"
                output_files[filename] = xlsx_bytes

    return output_files, no_data

@login_required
def index(request):
    # FASE 0: resetear en GET
    if request.method == 'GET':
        request.session.pop('uploaded_file_path', None)

    # Recuperamos el nombre del archivo subido, si existe
    uploaded_file_name = request.session.get('uploaded_file_name', '')

    upload_form = UploadFileForm(request.POST or None, request.FILES or None)

    # FASE 1: SUBIDA DE EXCEL con validación de .xlsx
    if request.method=='POST' and 'upload' in request.POST and upload_form.is_valid():
        f = upload_form.cleaned_data['file']
        if not f.name.lower().endswith('.xlsx'):
            messages.error(request, "Formato inválido: sólo se permiten archivos .xlsx.")
            return render(request, 'consignaciones_atico/index.html', {
                'upload_form':    upload_form,
                'formset':        ContactInfoFormSet(initial=[]),
                'editorial_list': [],
                'uploaded_file_name': '',
            })

        # 1️⃣ guardamos el .xlsx y su nombre en sesión
        temp_path = default_storage.save('temp/'+f.name, ContentFile(f.read()))
        request.session['uploaded_file_path'] = temp_path
        request.session['uploaded_file_name'] = f.name

        # reconstruir fase 2
        bytes_         = default_storage.open(temp_path).read()
        editorial_list = extract_editoriales(bytes_)
        initial = []
        saved   = load_contact_data()
        for ed in editorial_list:
            d = saved.get(ed, {}).copy()
            d['editorial'] = ed
            initial.append(d)
        formset = ContactInfoFormSet(initial=initial)

        return render(request, 'consignaciones_atico/index.html', {
            'upload_form':         upload_form,
            'formset':             formset,
            'editorial_list':      editorial_list,
            'uploaded_file_name':  request.session.get('uploaded_file_name', ''),
        })

    # siguientes fases: usamos el archivo en sesión si existe
    stored = request.session.get('uploaded_file_path')
    editorial_list = []
    if stored and default_storage.exists(stored):
        bytes_ = default_storage.open(stored).read()
        editorial_list = extract_editoriales(bytes_)

    initial = []
    saved   = load_contact_data()
    for ed in editorial_list:
        d = saved.get(ed, {}).copy()
        d['editorial'] = ed
        initial.append(d)
    formset = ContactInfoFormSet(request.POST or None, initial=initial)

    # FASE 2: guardar contactos
    if request.method=='POST' and 'save_contacts' in request.POST:
        if formset.is_valid():
            newd = {}
            for frm in formset:
                cd = frm.cleaned_data
                ed = cd.pop('editorial')
                newd[ed] = cd
            save_contact_data(newd)
            messages.success(request, "Contactos guardados correctamente.")
        else:
            messages.error(request, "Corrige los errores de contacto antes de guardar.")
        return render(request, 'consignaciones_atico/index.html', {
            'upload_form':         upload_form,
            'formset':             formset,
            'editorial_list':      editorial_list,
            'uploaded_file_name':  uploaded_file_name,
        })

    # FASE 3: generar ZIP
    if request.method=='POST' and 'generate_liquidaciones' in request.POST:
        if not stored or not default_storage.exists(stored):
            messages.error(request, "Debes procesar primero el archivo.")
        elif not formset.is_valid():
            messages.error(request, "Corrige los errores de contacto antes de generar.")
        else:
            ci = {
                frm.cleaned_data['editorial']: {
                    'PROVEEDOR':   frm.cleaned_data['PROVEEDOR'],
                    'CONTACTO':    frm.cleaned_data['CONTACTO'],
                    'FONO / MAIL': frm.cleaned_data['FONO_MAIL'],
                    'DESCUENTO':   frm.cleaned_data['DESCUENTO'],
                    'PAGO':        frm.cleaned_data['PAGO'],
                    'FECHA':       frm.cleaned_data['FECHA'],
                }
                for frm in formset
            }
            logo, no_data = load_logo_bytes(), []
            files, no_data = process_master_file(bytes_, logo, ci)

            if not files:
                messages.error(request, "No se generaron liquidaciones.")
                return render(request, 'consignaciones_atico/index.html', {
                    'upload_form':    upload_form,
                    'formset':        formset,
                    'editorial_list': editorial_list,
                    'uploaded_file_name': uploaded_file_name,
                })

            # — empaquetamos en memoria —
            zip_buf = BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zp:
                for fname, data in files.items():
                    zp.writestr(fname, data)
            zip_buf.seek(0)

            # codificamos a Base64 para incrustar en el template
            zip_b64 = base64.b64encode(zip_buf.read()).decode('ascii')
            generated_files = list(files.keys())

            messages.success(request, "Liquidaciones generadas correctamente.")
            # volvemos a renderizar el mismo index.html con TODO el contexto
            return render(request, 'consignaciones_atico/index.html', {
                'upload_form':          upload_form,
                'formset':              formset,
                'editorial_list':       editorial_list,
                'uploaded_file_name':   uploaded_file_name,
                # aquí va el feedback
                'generated_files':      generated_files,
                'zip_b64':              zip_b64,
                'no_data_editorials':   no_data,
            })

    # FASE 0 inicial o GET tras procesar
    return render(request, 'consignaciones_atico/index.html', {
        'upload_form':         upload_form,
        'formset':             formset,
        'editorial_list':      editorial_list,
        'uploaded_file_name':  uploaded_file_name,
    })
